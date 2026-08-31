"""WorkbookSpec → XLSX。"""

from __future__ import annotations

import math
from copy import copy
from pathlib import Path
from typing import Any
from unicodedata import east_asian_width

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.chart import BarChart, LineChart, Reference  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore[import-untyped]
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.cowork.artifact_renderers.contracts import (
    WorkbookCell,
    WorkbookChart,
    WorkbookSpec,
    WorksheetSpec,
)


def _apply_cell_style(cell: Any, spec: WorkbookCell) -> None:
    cell.font = Font(name="Aptos", size=10, color="17211D")
    cell.alignment = Alignment(vertical="top")
    if spec.style == "title":
        cell.font = Font(name="Aptos Display", size=18, bold=True, color="17211D")
    elif spec.style == "header":
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="167A5B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif spec.style == "metric":
        cell.font = Font(name="Aptos Display", size=20, bold=True, color="167A5B")
    elif spec.style == "currency":
        cell.number_format = "¥#,##0.00;[Red]-¥#,##0.00"
    elif spec.style == "percent":
        cell.number_format = "0.0%"
    elif spec.style == "date":
        cell.number_format = "yyyy-mm-dd"


def _display_width(value: object) -> int:
    text = "" if value is None else str(value)
    return max(
        (
            sum(2 if east_asian_width(character) in {"W", "F", "A"} else 1 for character in line)
            for line in text.splitlines() or [""]
        ),
        default=0,
    )


def _split_range(value: str, default_sheet: str) -> tuple[str, str]:
    if "!" not in value:
        return default_sheet, value
    sheet, _, cell_range = value.rpartition("!")
    return sheet.strip("'"), cell_range


def _reference(workbook: Workbook, value: str, default_sheet: str) -> Reference:
    sheet_name, cell_range = _split_range(value, default_sheet)
    worksheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return Reference(
        worksheet,
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
    )


def _add_chart(workbook: Workbook, worksheet: Worksheet, spec: WorkbookChart) -> None:
    chart = LineChart() if spec.chart_type == "line" else BarChart()
    if isinstance(chart, BarChart):
        chart.type = "bar" if spec.chart_type == "bar" else "col"
    chart.title = spec.title
    chart.style = 10
    chart.add_data(_reference(workbook, spec.data_range, worksheet.title), titles_from_data=True)
    chart.set_categories(_reference(workbook, spec.categories_range, worksheet.title))
    chart.height = 8
    chart.width = 14
    worksheet.add_chart(chart, spec.anchor)


def _render_sheet(workbook: Workbook, worksheet: Worksheet, spec: WorksheetSpec) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = spec.freeze_panes
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_view.zoomScale = 90
    for cell_spec in spec.cells:
        cell = worksheet[cell_spec.address]
        cell.value = cell_spec.formula if cell_spec.formula is not None else cell_spec.value
        _apply_cell_style(cell, cell_spec)
    for table_spec in spec.tables:
        start_col, start_row, _, _ = range_boundaries(f"{table_spec.anchor}:{table_spec.anchor}")
        for offset, header in enumerate(table_spec.headers):
            cell = worksheet.cell(row=start_row, column=start_col + offset, value=header)
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="167A5B")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_offset, values in enumerate(table_spec.rows, start=1):
            for column_offset, value in enumerate(values):
                worksheet.cell(
                    row=start_row + row_offset,
                    column=start_col + column_offset,
                    value=value,
                )
        end_row = start_row + max(1, len(table_spec.rows))
        end_col = start_col + len(table_spec.headers) - 1
        table = Table(
            displayName=table_spec.name,
            ref=f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    for chart_spec in spec.charts:
        _add_chart(workbook, worksheet, chart_spec)
    column_widths: dict[int, float] = {}
    for column in range(1, min(worksheet.max_column, 100) + 1):
        values = [
            worksheet.cell(row=row, column=column).value
            for row in range(1, min(worksheet.max_row, 500) + 1)
        ]
        longest = max((_display_width(value) for value in values), default=8)
        width = min(48.0, max(8.0, float(longest + 2)))
        column_widths[column] = width
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(1, min(worksheet.max_row, 10_000) + 1):
        required_lines = 1
        for column in range(1, min(worksheet.max_column, 100) + 1):
            cell = worksheet.cell(row=row, column=column)
            if cell.value is None:
                continue
            width = column_widths.get(column, 10.0)
            display = _display_width(cell.value)
            explicit_lines = len(str(cell.value).splitlines())
            lines = max(explicit_lines, math.ceil(display / max(1.0, width - 1)))
            if lines > 1:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment
            elif cell.alignment.vertical is None:
                alignment = copy(cell.alignment)
                alignment.vertical = "top"
                cell.alignment = alignment
            required_lines = max(required_lines, min(lines, 12))
        if required_lines > 1:
            worksheet.row_dimensions[row].height = max(15, 15 * required_lines)
    if worksheet.max_column > 8:
        worksheet.page_setup.orientation = "landscape"


def render_workbook(spec: WorkbookSpec, target: Path) -> None:
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    for index, sheet_spec in enumerate(spec.sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = sheet_spec.name
        _render_sheet(workbook, worksheet, sheet_spec)
    workbook.properties.title = spec.title
    workbook.properties.subject = spec.purpose or ""
    workbook.properties.creator = "WorkPilot"
    workbook.save(target)


__all__ = ["render_workbook"]
